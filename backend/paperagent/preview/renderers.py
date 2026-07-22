from __future__ import annotations

import csv
import json
import mimetypes
import subprocess
import sys
import zipfile
from dataclasses import dataclass
from pathlib import Path
from typing import Protocol

from openpyxl import load_workbook
from PIL import Image
from pypdf import PdfReader

from paperagent.ingestion.parsers import default_registry
from paperagent.preview.schemas import PreviewAnchor, PreviewFidelity, PreviewPart


@dataclass(frozen=True)
class RenderResult:
    fidelity: PreviewFidelity
    media_type: str
    payload: dict[str, object]
    parts: list[PreviewPart]
    capabilities: list[str]


class PreviewRenderer(Protocol):
    @property
    def name(self) -> str: ...

    @property
    def version(self) -> str: ...

    @property
    def extensions(self) -> tuple[str, ...]: ...

    def render(self, path: Path, *, file_id: str, source_hash: str) -> RenderResult: ...


def anchor(file_id: str, source_hash: str, format_name: str, **locator: object) -> PreviewAnchor:
    return PreviewAnchor.model_validate(
        {
            "source_file_id": file_id,
            "source_hash": source_hash,
            "format": format_name,
            **locator,
        }
    )


class PdfRenderer:
    name = "pdf"
    version = "1.0"
    extensions = (".pdf",)

    def render(self, path: Path, *, file_id: str, source_hash: str) -> RenderResult:
        reader = PdfReader(path)
        if reader.is_encrypted:
            return RenderResult(
                PreviewFidelity.METADATA,
                "application/pdf",
                {"encrypted": True, "page_count": None, "reason": "Password required"},
                [],
                ["system_open"],
            )
        parts = [
            PreviewPart(
                index=index - 1,
                kind="pdf_page",
                label=f"Page {index}",
                payload={"text": page.extract_text() or "", "rotation": page.rotation},
                anchor=anchor(
                    file_id,
                    source_hash,
                    "pdf",
                    page=index,
                    bbox=(0.0, 0.0, float(page.mediabox.width), float(page.mediabox.height)),
                ),
            )
            for index, page in enumerate(reader.pages, start=1)
        ]
        return RenderResult(
            PreviewFidelity.NATIVE,
            "application/pdf",
            {"page_count": len(parts), "initial_part_limit": 5},
            parts,
            ["search", "zoom", "page", "select", "annotate", "system_open"],
        )


class ImageRenderer:
    name = "image"
    version = "1.0"
    extensions = (".png", ".jpg", ".jpeg", ".webp", ".gif")

    def render(self, path: Path, *, file_id: str, source_hash: str) -> RenderResult:
        with Image.open(path) as image:
            payload = {
                "width": image.width,
                "height": image.height,
                "mode": image.mode,
                "format": image.format,
                "frames": getattr(image, "n_frames", 1),
            }
        part = PreviewPart(
            index=0,
            kind="image_surface",
            label="Image",
            payload={"width": payload["width"], "height": payload["height"]},
            anchor=anchor(file_id, source_hash, path.suffix, json_path="$image"),
        )
        return RenderResult(
            PreviewFidelity.NATIVE,
            mimetypes.guess_type(path.name)[0] or "image/*",
            payload,
            [part],
            ["zoom", "pan", "annotate", "ocr_layer", "system_open"],
        )


class StructuredDocumentRenderer:
    name = "structured-document"
    version = "1.0"
    extensions = (
        ".docx",
        ".pptx",
        ".md",
        ".markdown",
        ".txt",
        ".eml",
        ".mbox",
        ".msg",
        ".bib",
        ".ris",
        ".xml",
        ".json",
        ".ipynb",
    )

    def render(self, path: Path, *, file_id: str, source_hash: str) -> RenderResult:
        report = default_registry().import_file(path)
        parts: list[PreviewPart] = []
        for index, chunk in enumerate(report.source.chunks):
            locator = chunk.locator.model_dump(exclude_none=True)
            preview_anchor = self._anchor(file_id, source_hash, path.suffix, locator)
            parts.append(
                PreviewPart(
                    index=index,
                    kind=chunk.kind,
                    label=f"{chunk.kind} {index + 1}",
                    payload={"text": chunk.text, "metadata": chunk.metadata},
                    anchor=preview_anchor,
                )
            )
        media_type = report.source.media_type
        return RenderResult(
            PreviewFidelity.STRUCTURED,
            media_type,
            {"warnings": report.warnings, "structured_parts": len(parts)},
            parts,
            ["search", "select", "annotate", "system_open"],
        )

    @staticmethod
    def _anchor(
        file_id: str, source_hash: str, extension: str, locator: dict[str, object]
    ) -> PreviewAnchor:
        if "page" in locator:
            return anchor(file_id, source_hash, extension, page=locator["page"])
        if "slide" in locator:
            return anchor(file_id, source_hash, extension, slide=locator["slide"])
        if "sheet" in locator:
            return anchor(
                file_id,
                source_hash,
                extension,
                sheet=locator["sheet"],
                cell_range=locator.get("cell_range"),
            )
        if "line_start" in locator:
            return anchor(
                file_id,
                source_hash,
                extension,
                line_start=locator["line_start"],
                line_end=locator.get("line_end"),
            )
        if "message_id" in locator:
            return anchor(file_id, source_hash, extension, message_id=locator["message_id"])
        return anchor(
            file_id,
            source_hash,
            extension,
            json_path=locator.get("json_path") or "$",
        )


class TableRenderer:
    name = "table"
    version = "1.0"
    extensions = (".csv", ".tsv", ".xlsx", ".xlsm")

    def render(self, path: Path, *, file_id: str, source_hash: str) -> RenderResult:
        parts: list[PreviewPart] = []
        sheets: list[str] = []
        if path.suffix in {".csv", ".tsv"}:
            delimiter = "\t" if path.suffix == ".tsv" else ","
            with path.open(encoding="utf-8-sig", errors="replace", newline="") as stream:
                for row_index, row in enumerate(csv.reader(stream, delimiter=delimiter), start=1):
                    parts.append(
                        PreviewPart(
                            index=row_index - 1,
                            kind="table_row",
                            label=f"Row {row_index}",
                            payload={"cells": row},
                            anchor=anchor(
                                file_id,
                                source_hash,
                                path.suffix,
                                sheet="Sheet1",
                                cell_range=f"A{row_index}",
                            ),
                        )
                    )
            sheets = ["Sheet1"]
        else:
            workbook = load_workbook(path, read_only=True, data_only=False, keep_vba=False)
            index = 0
            for sheet in workbook.worksheets:
                sheets.append(sheet.title)
                for row in sheet.iter_rows():
                    cells = ["" if cell.value is None else str(cell.value) for cell in row]
                    if not any(cells):
                        continue
                    parts.append(
                        PreviewPart(
                            index=index,
                            kind="table_row",
                            label=f"{sheet.title}!{row[0].row}",
                            payload={"cells": cells},
                            anchor=anchor(
                                file_id,
                                source_hash,
                                path.suffix,
                                sheet=sheet.title,
                                cell_range=f"A{row[0].row}",
                            ),
                        )
                    )
                    index += 1
            workbook.close()
        return RenderResult(
            PreviewFidelity.STRUCTURED,
            mimetypes.guess_type(path.name)[0] or "text/csv",
            {"sheets": sheets, "row_count": len(parts), "virtualized": True},
            parts,
            ["search", "sheet", "cell_range", "select", "annotate", "system_open"],
        )


class CodeRenderer:
    name = "code-structured"
    version = "1.0"
    extensions = (
        ".py",
        ".js",
        ".ts",
        ".tsx",
        ".java",
        ".c",
        ".cpp",
        ".rs",
        ".go",
        ".yaml",
        ".yml",
    )

    def render(self, path: Path, *, file_id: str, source_hash: str) -> RenderResult:
        lines = path.read_text(encoding="utf-8", errors="replace").splitlines()
        parts = [
            PreviewPart(
                index=index - 1,
                kind="code_line",
                label=str(index),
                payload={"text": line},
                anchor=anchor(file_id, source_hash, path.suffix, line_start=index, line_end=index),
            )
            for index, line in enumerate(lines, start=1)
        ]
        return RenderResult(
            PreviewFidelity.STRUCTURED,
            "text/plain",
            {"language": path.suffix.lstrip("."), "line_count": len(lines), "read_only": True},
            parts,
            ["search", "line", "symbol", "select", "annotate", "system_open"],
        )


class HtmlSvgRenderer:
    name = "sanitized-active-content"
    version = "1.0"
    extensions = (".html", ".htm", ".svg")

    def render(self, path: Path, *, file_id: str, source_hash: str) -> RenderResult:
        worker = Path(__file__).with_name("_sanitize_worker.py")
        completed = subprocess.run(
            [sys.executable, "-I", str(worker), str(path)],
            check=True,
            capture_output=True,
            text=True,
            encoding="utf-8",
            timeout=10,
            creationflags=getattr(subprocess, "CREATE_NO_WINDOW", 0),
        )
        isolated = json.loads(completed.stdout)
        sanitized = str(isolated["html"])
        text = str(isolated["text"])
        part = PreviewPart(
            index=0,
            kind="sanitized_html",
            label="Safe content",
            payload={"html": sanitized, "text": text},
            anchor=anchor(file_id, source_hash, path.suffix, json_path="$sanitized"),
        )
        return RenderResult(
            PreviewFidelity.RENDERED,
            "image/svg+xml" if path.suffix == ".svg" else "text/html",
            {
                "active_content_removed": bool(isolated["changed"]),
                "process_isolated": True,
                "network_access": "not granted",
            },
            [part],
            ["search", "select", "annotate", "system_open"],
        )


class ArchiveRenderer:
    name = "archive-tree"
    version = "1.0"
    extensions = (".zip",)

    def render(self, path: Path, *, file_id: str, source_hash: str) -> RenderResult:
        parts: list[PreviewPart] = []
        with zipfile.ZipFile(path) as archive:
            for index, info in enumerate(archive.infolist()):
                member = Path(info.filename)
                if member.is_absolute() or ".." in member.parts:
                    raise ValueError("Archive contains unsafe path")
                parts.append(
                    PreviewPart(
                        index=index,
                        kind="archive_entry",
                        label=info.filename,
                        payload={"path": info.filename, "size": info.file_size},
                        anchor=anchor(file_id, source_hash, ".zip", json_path=f"entries[{index}]"),
                    )
                )
        return RenderResult(
            PreviewFidelity.STRUCTURED,
            "application/zip",
            {"entry_count": len(parts), "extraction_disabled": True},
            parts,
            ["search", "tree", "select", "system_open"],
        )


class MetadataRenderer:
    name = "safe-metadata"
    version = "1.0"
    extensions = ()

    def render(self, path: Path, *, file_id: str, source_hash: str) -> RenderResult:
        return RenderResult(
            PreviewFidelity.METADATA,
            mimetypes.guess_type(path.name)[0] or "application/octet-stream",
            {
                "name": path.name,
                "size_bytes": path.stat().st_size,
                "extension": path.suffix,
                "reason": "No safe in-app renderer is registered",
            },
            [],
            ["system_open"],
        )


DEFAULT_RENDERERS: tuple[PreviewRenderer, ...] = (
    PdfRenderer(),
    ImageRenderer(),
    StructuredDocumentRenderer(),
    TableRenderer(),
    CodeRenderer(),
    HtmlSvgRenderer(),
    ArchiveRenderer(),
)
